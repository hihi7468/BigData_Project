# cls 기능 사용 위해 import
import os                                   
# sleep 기능 사용 위해 import
from time import sleep                     
# 엑셀저장기능 활용위해 import
import openpyxl

# 키오스크 클래스
class kiosk():
    menu_type_check = 0
    menu_name_check = 0
    def __init__(self):
        self.order_list = {} 
        self.order_id = 0
        self.order_count = 0         
        self.menu_list = {'라면':[3000,'분식', 10],'떡볶이':[4000,'분식', 10],'된장찌개':[7000,'한식', 10],'비빔밥':[7000,'한식', 10]
                          ,'짜장면':[5000,'중식',10]}
        self.menu_type_list=['분식', '한식','중식']
        self.sum_type_list={'분식': [0, self.order_count], '한식': [0 , self.order_count],'중식': [0 , self.order_count]}
        self.sum_menu_list={'라면':[0, self.order_count],'된장찌개':[0, self.order_count],'떡볶이':[0, self.order_count],
                            '비빔밥':[0, self.order_count],'짜장면':[0, self.order_count]}
        self.print_status()
# 실제 적용시 좀더 추가하려했던 메뉴들
# 메뉴 전체    {'라면':[3000,'분식', 100], '떡볶이':[3000,'분식',100],'우동':[4000,'분식',100],'김밥':[2000,'분식', 100],
#             '소고기덮밥':[6000,'분식',100], '제육덮밥':[6000,'분식',100],'오므라이스':[6000,'분식',100], 
#             '된장찌개':[7000,'한식', 100], '순두부찌개':[7000,'한식',100],'김치찌개':[7000,'한식',100],
#             '백반정식':[7000,'한식',100], '냉면':[7000,'한식',100], '돌솥밥':[7000,'한식',100], '제육볶음':[7000,'한식',100],
#             '짜장면':[5000,'중식',100],'간짜장':[6000,'중식',100],'짬뽕':[6000,'중식',100],'고기짬뽕':[6000,'중식',100],
#             '탕수육':[15000,'중식',100],'볶음밥':[6000,'중식',100],'세트A':[13000,'중식',100],'세트B':[14000,'중식',100],
#             '돈까스':[7000,'양식',100],'필라프':[7000,'양식',100],'미트스파게티':[7000,'양식',100],'파스타':[8000,'양식',100]}


    # 최초 메뉴 팝업창
    def print_status(self):
        while(True):
            os.system('cls')
            key = 0
            print("원하시는 번호를 선택해주세요\n")
            print(" ① : 메뉴 선택창 \n ② : 결제창 \n  관리자전용창:??")
            key = int(input("번호 입력:"))
            if key == 1:                            #소비자 메뉴 선택창
                self.show_menu()
            elif key == 2:                          #소비자 결제창
                self.purchase()
            elif key == 99:                         #관리자 창 - 암호입력시 진입
                self.admin_view()

    # 소비자 메뉴 선택창 
    def show_menu(self):
        os.system('cls')
        self.menu_name_check = 0
        print("메뉴의 종류를 골라주세요.\n")
        for i in range(1,len(self.menu_type_list)+1):
            print("%d번:%s "%(i,self.menu_type_list[i-1]), end ="")
        print("\n")
        try:
            menu_number = int(input("번호:"))
        except:
            print("숫자만 넣어주세요.\n")
            self.show_menu()
        if menu_number > len(self.menu_type_list) or menu_number == 0:
            print("해당 번호는 존재하지 않습니다. 다시 입력해주세요.\n")
            self.show_menu()
        else:
            print("%d번 %s의 메뉴 리스트입니다.\n" %(menu_number, self.menu_type_list[menu_number-1]))
            for i in range(0, len(self.menu_list) ):
                if (list(self.menu_list['%s' %(list(self.menu_list.keys())[i])])[1]) == self.menu_type_list[menu_number-1]:
                    print("%s의 가격은"%(list(self.menu_list.keys())[i]), end="")
                    print(list(self.menu_list['%s' %(list(self.menu_list.keys())[i])])[0], end="")
                    print("입니다.")
            try:
                m_n= input("원하시는 메뉴를 입력해주세요.\n")
            except:
                print("잘못된 값을 입력하였습니다. 다시 시도해주세요.\n")
                sleep(2)
                self.show_menu()
            for i in range(0, len(self.menu_list)):
                if(m_n ==list(self.menu_list.keys())[i]):
                    self.menu_name_check +=1
            if self.menu_name_check != 1:
                print("해당되는 메뉴가 없습니다.\n")
                sleep(2)
                self.show_menu()
            elif (self.menu_name_check == 1) and (self.menu_list[m_n][2] != 0) :
                print("%s를 주문하시겠습니까?.\n"%(m_n))
                print("이름과 전화번호를 적어주세요.\n")
                s_name = input("이름:")
                s_tel = input("전화번호:")
                self.push_list(s_name, s_tel, m_n)
            elif self.menu_list[m_n][2] <= 0:
                print("해당 메뉴의 재고가 전부 소진되었습니다. 다른 메뉴를 선택해주세요.\n")
                sleep(1)
                self.show_menu()

    # 소비자 메뉴추가과정
    def push_list(self, user_name, user_tel, menu_name):
        os.system('cls')
        self.order_id += 1
        self.order_list[self.order_id] = [user_name, user_tel, menu_name, 0]
        print(self.order_list)
        print(self.order_list.keys())

    # 소비자 결제창
    def purchase(self):
        os.system('cls')
        total_price = 0
        cash = 0
        purchase_name = input("이름을 입력하세요\n")
        for i in range (1, self.order_id+1):
            if self.order_list[i][3] == 0:
                if self.order_list[i][0] == purchase_name:  
                    total_price += self.menu_list[self.order_list[i][2]][0] 
                elif self.order_list[i][3] == 1:
                    print("결제처리가 완료된 번호입니다.\n")
        sleep(1)
        print("총 금액은 %d원 입니다." %total_price)
        while cash < total_price:
            cash = int(input("투입 금액을 입력하세요"))
            if total_price > cash:
                while(total_price>cash):
                    print("잔액이 부족합니다.")
                    print("현재 투입금액은%d입니다.\n"%cash)
                    print("%d원이 부족합니다."%(total_price-cash))
                    cash += int(input("투입 금액을 입력하세요"))
                    if(total_price<= cash):
                        print("결제를 완료했습니다. 거스름돈은 %d원 입니다" %(cash - total_price))
                        sleep(2)
            else:
                print("결제를 완료했습니다. 거스름돈은 %d원 입니다" %(cash - total_price))
                sleep(2)
        for i in range (1, self.order_id+1):
            if purchase_name == self.order_list[i][0]:
                 self.order_list[i][3] = 1
                 self.menu_list[self.order_list[i][2]][2] -= 1

      
 
    # 관리자 선택창    
    def admin_view(self):
        p_key = 0
        os.system('cls')
        while(p_key != 99):
            os.system('cls')
            print("관리자 창입니다.\n 관리할 사항을 눌러주세요.\n")
            print(" ①번: 추가 메뉴 입력 \n ②번: 기존 메뉴 삭제 \n ③번: 매출 확인 \n ④번: 매출 엑셀로 저장 \n ⑨⑨번: 나가기")
            p_key = int(input("번호 입력:"))                    #암호입력
            if p_key == 3:                                      #매출확인
                self.count_sell()
            elif p_key == 2:
                print("삭제할 메뉴의 이름을 적어주세요")          #메뉴삭제
                d_name = input("메뉴의 이름:")
                self.delete_menu(d_name)
            elif p_key == 1:                                    #메뉴추가
                print("추가할 메뉴의 이름, 가격, 종류, 추가할 재고량을 적어주세요.\n")
                s_name=input("추가 메뉴명:")
                s_price = input("추가 메뉴 가격:")
                s_type = input("추가 메뉴 종류:")
                s_amount = input("추가할 재고량:")
                self.add_menu(s_name, s_price, s_type, s_amount)
            elif p_key == 4:                                    #엑셀파일저장
                self.write_sell()
                print("엑셀로 저장을 완료합니다\n")
                sleep(1)


    # 관리자 창의 메뉴추가창
    def add_menu(self, menu_name, menu_price, menu_type, menu_amount):
        os.system('cls')
        self.menu_list[menu_name] =[menu_price, menu_type, menu_amount]
        self.sum_menu_list[menu_name] = [0, self.order_count]
        self.menu_type_check = 0
        for i in range(0, len(self.menu_type_list)):
            if self.menu_type_list[i] == menu_type :
                self.menu_type_check += 1
                break
        if self.menu_type_check != 1:
            self.menu_type_list += [menu_type]
            self.sum_type_list[menu_type]= [0, self.order_count]

    # 관리자 창의 메뉴삭제창
    def delete_menu(self, menu_name):
        os.system('cls')
        del self.menu_list[menu_name]

    # 관리자 창의 매출확인
    def count_sell(self):
        os.system('cls')
        for i in range(1, len(self.order_list)+1):
            self.sum_menu_list[self.order_list[i][2]][0] = 0
            self.sum_menu_list[self.order_list[i][2]][1] = 0
            self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][0]= 0
            self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][1]= 0
        print("매출을 체크합니다.\n")
        for i in range(1, len(self.order_list)+1):
            if self.order_list[i][3] == 1:
                self.sum_menu_list[self.order_list[i][2]][0] += self.menu_list[self.order_list[i][2]][0]
                self.sum_menu_list[self.order_list[i][2]][1] += 1
                self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][0] += self.menu_list[self.order_list[i][2]][0]
                self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][1] += 1

        for i in range(0, len(self.sum_menu_list)):         #매뉴 개당 판매 출력 
            print(list(self.sum_menu_list.keys())[i])
            print("메뉴의 총 판매 금액은 %d 원입니다.\n"%self.sum_menu_list[list(self.sum_menu_list.keys())[i]][0])
            print("총 판매량은 %d개 입니다. \n"%self.sum_menu_list[list(self.sum_menu_list.keys())[i]][1])
            sleep(2)
        for i in range(0, len(self.sum_type_list)):         #메뉴 종류별 판매 출력
            print(list(self.sum_type_list.keys())[i])
            print("메뉴류의 총 판매 금액은 %d 원입니다.\n" %self.sum_type_list[list(self.sum_type_list.keys())[i]][0])
            print("총 판매량은 %d개 입니다. \n" %self.sum_type_list[list(self.sum_type_list.keys())[i]][1])
            sleep(2)


    # 엑셀 파일로 저장                    
    def write_sell(self):
        for i in range(1, len(self.order_list)+1):      # 엑셀로 저장시 디폴트값 
            self.sum_menu_list[self.order_list[i][2]][0] = 0
            self.sum_menu_list[self.order_list[i][2]][1] = 0
            self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][0]= 0
            self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][1]= 0
        for i in range(1, len(self.order_list)+1):
            if self.order_list[i][3] == 1:              # 추가값을 연산
                self.sum_menu_list[self.order_list[i][2]][0] += self.menu_list[self.order_list[i][2]][0]
                self.sum_menu_list[self.order_list[i][2]][1] += 1
                self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][0] += self.menu_list[self.order_list[i][2]][0]
                self.sum_type_list[self.menu_list[self.order_list[i][2]][1]][1] += 1

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'test sheet'
        ws['A1'] = '메뉴 리스트'
        ws['B1'] = '총 판매액'
        ws['C1'] = '총 판매 개수'
        ws['E1'] = '메뉴 타입'
        ws['F1'] = '총 판매액'
        ws['G1'] = '총 판매량'

        for i in range(0, len(self.sum_menu_list)):             #총 판매액
            ws['A%d'%(i+2)] = list(self.sum_menu_list.keys())[i]
            ws['B%d'%(i+2)] = self.sum_menu_list[list(self.sum_menu_list.keys())[i]][0]
            ws['C%d'%(i+2)] = self.sum_menu_list[list(self.sum_menu_list.keys())[i]][1]
        for i in range(0, len(self.sum_type_list)):             #총 판매량
            ws['E%d'%(i+2)] = list(self.sum_type_list.keys())[i]
            ws['F%d'%(i+2)] = self.sum_type_list[list(self.sum_type_list.keys())[i]][0]
            ws['G%d'%(i+2)] = self.sum_type_list[list(self.sum_type_list.keys())[i]][1]

        wb.save(r'C:\Users\Admin\Documents\카카오톡 받은 파일\cell.xlsx')
        wb.close()

    
a = kiosk()
#출력
